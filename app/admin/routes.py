import os
import shutil
from datetime import datetime, timedelta
from types import SimpleNamespace
from flask import render_template, redirect, url_for, flash, request, current_app, jsonify
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from app.admin import admin
from app.extensions import db
from app.models.user import User
from app.models.cms import Page, Post, Event, Download, Category, ContactMessage, Testimonial, HomeSection, Album, Media
from app.models.site import Setting
from app.forms.cms import PageForm, PostForm, EventForm, DownloadForm, CategoryForm
from app.forms.site import SettingsForm, TestimonialForm

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('You do not have permission to access this page.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function

def format_file_size(size_bytes):
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"

# =========================================================================
# DASHBOARD
# =========================================================================
@admin.route('/')
@admin.route('/dashboard')
@login_required
@admin_required
def dashboard():
    stats = {
        'pages': Page.query.count(),
        'posts': Post.query.count(),
        'events': Event.query.count(),
        'downloads': Download.query.count(),
        'users': User.query.count(),
        'messages': ContactMessage.query.filter_by(is_read=False).count(),
        'testimonials': Testimonial.query.count(),
        'categories': Category.query.count(),
    }
    return render_template('admin/dashboard.html', stats=stats, recent_logs=[], notifications=[])

@admin.route('/api/analytics/dashboard-metrics')
@login_required
@admin_required
def dashboard_metrics():
    today = datetime.utcnow().date()
    labels = [
        (today - timedelta(days=offset)).strftime('%b ') + str((today - timedelta(days=offset)).day)
        for offset in range(6, -1, -1)
    ]
    return jsonify({'labels': labels, 'views': [0] * len(labels)})

@admin.route('/api/sections/reorder', methods=['POST'])
@login_required
@admin_required
def reorder_sections():
    payload = request.get_json(silent=True) or {}
    order_ids = payload.get('order')
    if not isinstance(order_ids, list) or not all(isinstance(item, int) for item in order_ids):
        return jsonify({'success': False, 'message': 'Invalid section order.'}), 400

    sections = HomeSection.query.all()
    sections_by_id = {section.id: section for section in sections}
    if set(order_ids) != set(sections_by_id) or len(order_ids) != len(sections_by_id):
        return jsonify({'success': False, 'message': 'Section order does not match current sections.'}), 400

    for position, section_id in enumerate(order_ids):
        sections_by_id[section_id].order = position
    db.session.commit()
    return jsonify({'success': True})

# =========================================================================
# PAGE MANAGEMENT
# =========================================================================
@admin.route('/pages')
@login_required
@admin_required
def list_pages():
    pages = Page.query.order_by(Page.updated_at.desc()).all()
    return render_template('admin/pages_list.html', pages=pages)

@admin.route('/pages/new', methods=['GET', 'POST'])
@login_required
@admin_required
def new_page():
    form = PageForm()
    if form.validate_on_submit():
        existing = Page.query.filter_by(slug=form.slug.data).first()
        if existing:
            flash('A page with this URL slug already exists.', 'danger')
            return render_template('admin/page_form.html', form=form, title="Create New Page")
            
        page_item = Page(
            title=form.title.data,
            slug=form.slug.data,
            content=form.content.data,
            seo_title=form.seo_title.data,
            seo_description=form.seo_description.data,
            is_published=form.is_published.data
        )
        db.session.add(page_item)
        db.session.commit()
        flash('Page created successfully.', 'success')
        return redirect(url_for('admin.list_pages'))
    return render_template('admin/page_form.html', form=form, title="Create New Page")

@admin.route('/pages/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_page(id):
    page_item = Page.query.get_or_404(id)
    form = PageForm(obj=page_item)
    if form.validate_on_submit():
        existing = Page.query.filter_by(slug=form.slug.data).first()
        if existing and existing.id != id:
            flash('A page with this URL slug already exists.', 'danger')
            return render_template('admin/page_form.html', form=form, title="Edit Page")
            
        page_item.title = form.title.data
        page_item.slug = form.slug.data
        page_item.content = form.content.data
        page_item.seo_title = form.seo_title.data
        page_item.seo_description = form.seo_description.data
        page_item.is_published = form.is_published.data
        db.session.commit()
        flash('Page updated successfully.', 'success')
        return redirect(url_for('admin.list_pages'))
    return render_template('admin/page_form.html', form=form, title="Edit Page")

@admin.route('/pages/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_page(id):
    page_item = Page.query.get_or_404(id)
    db.session.delete(page_item)
    db.session.commit()
    flash('Page deleted successfully.', 'success')
    return redirect(url_for('admin.list_pages'))

# =========================================================================
# CATEGORY MANAGEMENT
# =========================================================================
@admin.route('/categories', methods=['GET', 'POST'])
@login_required
@admin_required
def list_categories():
    form = CategoryForm()
    if form.validate_on_submit():
        existing = Category.query.filter_by(slug=form.slug.data).first()
        if existing:
            flash('A category with this slug already exists.', 'danger')
        else:
            cat = Category(
                name=form.name.data,
                slug=form.slug.data,
                description=form.description.data
            )
            db.session.add(cat)
            db.session.commit()
            flash('Category created successfully.', 'success')
            return redirect(url_for('admin.list_categories'))
    categories = Category.query.order_by(Category.name.asc()).all()
    return render_template('admin/categories.html', categories=categories, form=form)

@admin.route('/categories/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_category(id):
    cat = Category.query.get_or_404(id)
    # unbind posts
    for post in cat.posts:
        post.category_id = None
    db.session.delete(cat)
    db.session.commit()
    flash('Category deleted successfully.', 'success')
    return redirect(url_for('admin.list_categories'))

# =========================================================================
# POST MANAGEMENT
# =========================================================================
@admin.route('/posts')
@login_required
@admin_required
def list_posts():
    posts = Post.query.order_by(Post.created_at.desc()).all()
    return render_template('admin/posts_list.html', posts=posts)

@admin.route('/posts/new', methods=['GET', 'POST'])
@login_required
@admin_required
def new_post():
    form = PostForm()
    categories = Category.query.order_by(Category.name.asc()).all()
    form.category_id.choices = [(0, '-- None / Uncategorized --')] + [(c.id, c.name) for c in categories]
    
    if form.validate_on_submit():
        existing = Post.query.filter_by(slug=form.slug.data).first()
        if existing:
            flash('A post with this URL slug already exists.', 'danger')
            return render_template('admin/post_form.html', form=form, title="Create New Post")
            
        featured_image_filename = None
        if form.featured_image.data:
            f = form.featured_image.data
            s_name = secure_filename(f.filename)
            ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
            featured_image_filename = f"post_{ts}_{s_name}"
            f.save(os.path.join(current_app.config['UPLOAD_FOLDER'], featured_image_filename))

        post_item = Post(
            title=form.title.data,
            slug=form.slug.data,
            content=form.content.data,
            excerpt=form.excerpt.data,
            tags=form.tags.data,
            featured_image=featured_image_filename,
            seo_title=form.seo_title.data,
            seo_description=form.seo_description.data,
            comments_enabled=form.comments_enabled.data,
            is_published=form.is_published.data,
            category_id=form.category_id.data if form.category_id.data != 0 else None,
            author_id=current_user.id
        )
        
        db.session.add(post_item)
        db.session.commit()
        flash('News post created successfully.', 'success')
        return redirect(url_for('admin.list_posts'))
        
    return render_template('admin/post_form.html', form=form, title="Create New Post")

@admin.route('/posts/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_post(id):
    post_item = Post.query.get_or_404(id)
    form = PostForm(obj=post_item)
    categories = Category.query.order_by(Category.name.asc()).all()
    form.category_id.choices = [(0, '-- None / Uncategorized --')] + [(c.id, c.name) for c in categories]
    
    if request.method == 'GET' and post_item.category_id:
        form.category_id.data = post_item.category_id

    if form.validate_on_submit():
        existing = Post.query.filter_by(slug=form.slug.data).first()
        if existing and existing.id != id:
            flash('A post with this URL slug already exists.', 'danger')
            return render_template('admin/post_form.html', form=form, title="Edit Post")
            
        if form.featured_image.data:
            f = form.featured_image.data
            s_name = secure_filename(f.filename)
            ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
            filename = f"post_{ts}_{s_name}"
            f.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
            post_item.featured_image = filename

        post_item.title = form.title.data
        post_item.slug = form.slug.data
        post_item.content = form.content.data
        post_item.excerpt = form.excerpt.data
        post_item.tags = form.tags.data
        post_item.seo_title = form.seo_title.data
        post_item.seo_description = form.seo_description.data
        post_item.comments_enabled = form.comments_enabled.data
        post_item.is_published = form.is_published.data
        post_item.category_id = form.category_id.data if form.category_id.data != 0 else None
        
        db.session.commit()
        flash('News post updated successfully.', 'success')
        return redirect(url_for('admin.list_posts'))
        
    return render_template('admin/post_form.html', form=form, title="Edit Post")

@admin.route('/posts/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_post(id):
    post_item = Post.query.get_or_404(id)
    db.session.delete(post_item)
    db.session.commit()
    flash('News post deleted successfully.', 'success')
    return redirect(url_for('admin.list_posts'))

# =========================================================================
# EVENT MANAGEMENT
# =========================================================================
@admin.route('/events')
@login_required
@admin_required
def list_events():
    events = Event.query.order_by(Event.start_time.desc()).all()
    return render_template('admin/events_list.html', events=events)

@admin.route('/events/new', methods=['GET', 'POST'])
@login_required
@admin_required
def new_event():
    form = EventForm()
    if form.validate_on_submit():
        existing = Event.query.filter_by(slug=form.slug.data).first()
        if existing:
            flash('An event with this URL slug already exists.', 'danger')
            return render_template('admin/event_form.html', form=form, title="Create New Event")
            
        featured_image_filename = None
        if form.featured_image.data:
            f = form.featured_image.data
            s_name = secure_filename(f.filename)
            ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
            featured_image_filename = f"evt_{ts}_{s_name}"
            f.save(os.path.join(current_app.config['UPLOAD_FOLDER'], featured_image_filename))

        evt = Event(
            title=form.title.data,
            slug=form.slug.data,
            description=form.description.data,
            start_time=form.start_time.data,
            end_time=form.end_time.data,
            venue=form.venue.data,
            organizer=form.organizer.data,
            featured_image=featured_image_filename,
            registration_url=form.registration_url.data,
            countdown_enabled=form.countdown_enabled.data,
            is_published=form.is_published.data
        )
        db.session.add(evt)
        db.session.commit()
        flash('Event created successfully.', 'success')
        return redirect(url_for('admin.list_events'))
        
    return render_template('admin/event_form.html', form=form, title="Create New Event")

@admin.route('/events/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_event(id):
    evt = Event.query.get_or_404(id)
    form = EventForm(obj=evt)
    if form.validate_on_submit():
        existing = Event.query.filter_by(slug=form.slug.data).first()
        if existing and existing.id != id:
            flash('An event with this URL slug already exists.', 'danger')
            return render_template('admin/event_form.html', form=form, title="Edit Event")
            
        if form.featured_image.data:
            f = form.featured_image.data
            s_name = secure_filename(f.filename)
            ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
            filename = f"evt_{ts}_{s_name}"
            f.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
            evt.featured_image = filename

        evt.title = form.title.data
        evt.slug = form.slug.data
        evt.description = form.description.data
        evt.start_time = form.start_time.data
        evt.end_time = form.end_time.data
        evt.venue = form.venue.data
        evt.organizer = form.organizer.data
        evt.registration_url = form.registration_url.data
        evt.countdown_enabled = form.countdown_enabled.data
        evt.is_published = form.is_published.data
        db.session.commit()
        flash('Event updated successfully.', 'success')
        return redirect(url_for('admin.list_events'))
        
    return render_template('admin/event_form.html', form=form, title="Edit Event")

@admin.route('/events/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_event(id):
    evt = Event.query.get_or_404(id)
    db.session.delete(evt)
    db.session.commit()
    flash('Event deleted successfully.', 'success')
    return redirect(url_for('admin.list_events'))

# =========================================================================
# DOWNLOAD CENTER
# =========================================================================
@admin.route('/downloads', methods=['GET', 'POST'])
@login_required
@admin_required
def list_downloads():
    downloads = Download.query.order_by(Download.created_at.desc()).all()
    form = DownloadForm()
    if form.validate_on_submit():
        if form.file.data:
            f = form.file.data
            safe_name = secure_filename(f.filename)
            timestamp = datetime.utcnow().strftime('%Y%m%d%H%M%S')
            filename = f"dl_{timestamp}_{safe_name}"
            
            filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            f.save(filepath)
            
            ext = os.path.splitext(safe_name)[1].replace('.', '').upper()
            
            dl = Download(
                title=form.title.data,
                filename=filename,
                file_type=ext,
                category=form.category.data,
                download_count=0
            )
            db.session.add(dl)
            db.session.commit()
            flash('File uploaded successfully.', 'success')
            return redirect(url_for('admin.list_downloads'))
            
    return render_template('admin/downloads.html', downloads=downloads, form=form)

@admin.route('/downloads/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_download(id):
    dl = Download.query.get_or_404(id)
    upload_folder = current_app.config['UPLOAD_FOLDER']
    filepath = os.path.join(upload_folder, dl.filename)
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
        except Exception:
            pass
            
    db.session.delete(dl)
    db.session.commit()
    flash('Download document deleted successfully.', 'success')
    return redirect(url_for('admin.list_downloads'))

# =========================================================================
# INBOX MESSAGES
# =========================================================================
@admin.route('/inbox')
@login_required
@admin_required
def inbox():
    messages = ContactMessage.query.order_by(ContactMessage.created_at.desc()).all()
    return render_template('admin/inbox.html', messages=messages)

@admin.route('/inbox/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def view_message(id):
    msg = ContactMessage.query.get_or_404(id)
    if not msg.is_read:
        msg.is_read = True
        db.session.commit()

    if request.method == 'POST':
        reply_content = request.form.get('reply_content')
        if reply_content:
            msg.reply_content = reply_content
            msg.replied_at = datetime.utcnow()
            db.session.commit()
            flash('Reply recorded successfully.', 'success')
            return redirect(url_for('admin.view_message', id=id))

    return render_template('admin/inbox_detail.html', msg=msg)

@admin.route('/inbox/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_message(id):
    msg = ContactMessage.query.get_or_404(id)
    db.session.delete(msg)
    db.session.commit()
    flash('Message deleted successfully.', 'success')
    return redirect(url_for('admin.inbox'))

# =========================================================================
# APPEARANCE & HOMEPAGE MANAGER
# =========================================================================
@admin.route('/appearance')
@login_required
@admin_required
def appearance():
    return redirect(url_for('admin.settings'))

@admin.route('/appearance/sections/<int:id>/edit', methods=['POST'])
@login_required
@admin_required
def edit_section(id):
    sec = HomeSection.query.get_or_404(id)
    sec.title = request.form.get('title')
    sec.subtitle = request.form.get('subtitle')
    sec.content = request.form.get('content')
    
    if sec.name == 'stats' and 'landing_image' in request.files:
        f = request.files['landing_image']
        if f and f.filename:
            s_name = secure_filename(f.filename)
            ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
            filename = f"landing_{ts}_{s_name}"
            f.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
            url = f"/static/uploads/{filename}"
            s = Setting.query.filter_by(key='landing_image_url').first()
            if not s:
                s = Setting(key='landing_image_url', value=url)
                db.session.add(s)
            else:
                s.value = url
                
    db.session.commit()
    flash(f"Section '{sec.name}' updated successfully.", 'success')
    return redirect(url_for('admin.appearance'))

@admin.route('/appearance/sections/<int:id>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_section(id):
    sec = HomeSection.query.get_or_404(id)
    sec.is_enabled = not sec.is_enabled
    db.session.commit()
    status = 'enabled' if sec.is_enabled else 'disabled'
    flash(f"Section '{sec.name}' has been {status}.", 'success')
    return redirect(url_for('admin.appearance'))

@admin.route('/appearance/testimonials/new', methods=['POST'])
@login_required
@admin_required
def new_testimonial():
    form = TestimonialForm()
    if form.validate_on_submit():
        photo_filename = None
        if form.photo.data:
            f = form.photo.data
            s_name = secure_filename(f.filename)
            ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
            photo_filename = f"testi_{ts}_{s_name}"
            f.save(os.path.join(current_app.config['UPLOAD_FOLDER'], photo_filename))

        testi = Testimonial(
            name=form.name.data,
            role=form.role.data,
            content=form.content.data,
            rating=form.rating.data,
            photo_path=photo_filename,
            is_published=True
        )
        db.session.add(testi)
        db.session.commit()
        flash('Testimonial review added successfully.', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{field}: {error}", 'danger')
    return redirect(url_for('admin.appearance'))

@admin.route('/appearance/testimonials/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_testimonial(id):
    testi = Testimonial.query.get_or_404(id)
    db.session.delete(testi)
    db.session.commit()
    flash('Testimonial removed successfully.', 'success')
    return redirect(url_for('admin.appearance'))

# =========================================================================
# MEDIA MANAGER
# =========================================================================
@admin.route('/upload')
@admin.route('/media')
@login_required
@admin_required
def list_media():
    upload_folder = current_app.config['UPLOAD_FOLDER']
    files_list = []
    if os.path.exists(upload_folder):
        for f in os.listdir(upload_folder):
            fp = os.path.join(upload_folder, f)
            if os.path.isfile(fp):
                size = os.path.getsize(fp)
                files_list.append(SimpleNamespace(
                    name=f,
                    url=url_for('static', filename='uploads/' + f),
                    size=format_file_size(size)
                ))
    return render_template('admin/media.html', files=files_list)

@admin.route('/media/upload', methods=['POST'])
@login_required
@admin_required
def upload_media():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    s_name = secure_filename(f.filename)
    ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
    filename = f"media_{ts}_{s_name}"
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    f.save(filepath)
    return jsonify({
        'success': True,
        'name': filename,
        'url': url_for('static', filename='uploads/' + filename),
        'size': format_file_size(os.path.getsize(filepath))
    })

@admin.route('/media/delete/<path:filename>', methods=['POST'])
@login_required
@admin_required
def delete_media(filename):
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
            flash(f"Media file '{filename}' deleted.", 'success')
        except Exception as e:
            flash(f"Error deleting file: {e}", 'danger')
    return redirect(url_for('admin.list_media'))

# =========================================================================
# SYSTEM SETTINGS
# =========================================================================
@admin.route('/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def settings():
    form = SettingsForm()
    
    fields = [
        'school_name', 'custom_domain', 'primary_color', 'secondary_color', 'footer_bg_color',
        'contact_email', 'contact_phone', 'contact_address', 'google_maps_url',
        'social_facebook', 'social_twitter', 'social_instagram', 'social_youtube', 'social_whatsapp',
        'seo_meta_description', 'analytics_code', 'footer_text'
    ]
    
    if request.method == 'GET':
        for field in fields:
            s = Setting.query.filter_by(key=field).first()
            if s and hasattr(form, field):
                getattr(form, field).data = s.value

    if form.validate_on_submit():
        for field in fields:
            val = getattr(form, field).data
            s = Setting.query.filter_by(key=field).first()
            if not s:
                s = Setting(key=field, value=val or '')
                db.session.add(s)
            else:
                s.value = val or ''

        # Handle file uploads
        file_fields = [
            ('logo', 'logo_url'),
            ('favicon', 'favicon_url'),
            ('hero_image', 'hero_image_url'),
            ('welcome_image', 'welcome_image_url'),
            ('principal_image', 'principal_image_url'),
        ]
        for form_file_field, setting_key in file_fields:
            file_data = getattr(form, form_file_field).data
            if file_data and hasattr(file_data, 'filename') and file_data.filename:
                s_name = secure_filename(file_data.filename)
                ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
                saved_name = f"set_{setting_key}_{ts}_{s_name}"
                file_data.save(os.path.join(current_app.config['UPLOAD_FOLDER'], saved_name))
                url_val = f"/static/uploads/{saved_name}"
                
                s = Setting.query.filter_by(key=setting_key).first()
                if not s:
                    s = Setting(key=setting_key, value=url_val)
                    db.session.add(s)
                else:
                    s.value = url_val

        db.session.commit()
        flash('Settings updated successfully.', 'success')
        return redirect(url_for('admin.settings'))

    sections = HomeSection.query.order_by(HomeSection.order.asc()).all()
    testimonials = Testimonial.query.order_by(Testimonial.created_at.desc()).all()
    test_form = TestimonialForm()
    return render_template(
        'admin/settings.html',
        form=form,
        sections=sections,
        testimonials=testimonials,
        test_form=test_form,
    )

# =========================================================================
# DATABASE BACKUPS
# =========================================================================
@admin.route('/backups')
@login_required
@admin_required
def backups():
    backup_dir = os.path.join(current_app.root_path, '..', 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    
    backup_list = []
    for f in os.listdir(backup_dir):
        fp = os.path.join(backup_dir, f)
        if os.path.isfile(fp):
            stat = os.stat(fp)
            backup_list.append(SimpleNamespace(
                filename=f,
                size=format_file_size(stat.st_size),
                created_at=datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
            ))
    backup_list.sort(key=lambda x: x.created_at, reverse=True)
    return render_template('admin/backups.html', backups=backup_list)

@admin.route('/backups/run', methods=['POST'])
@login_required
@admin_required
def run_backup():
    backup_dir = os.path.join(current_app.root_path, '..', 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    
    db_path = os.path.join(current_app.instance_path, 'school.db')
    if not os.path.exists(db_path):
        flash('Active database file not found for backup.', 'danger')
        return redirect(url_for('admin.backups'))
        
    ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    backup_filename = f"backup_{ts}.db"
    dest_path = os.path.join(backup_dir, backup_filename)
    
    shutil.copy2(db_path, dest_path)
    flash(f"Database backup '{backup_filename}' created successfully.", 'success')
    return redirect(url_for('admin.backups'))

@admin.route('/backups/<path:filename>/restore', methods=['POST'])
@login_required
@admin_required
def restore_backup(filename):
    backup_dir = os.path.join(current_app.root_path, '..', 'backups')
    src_path = os.path.join(backup_dir, filename)
    db_path = os.path.join(current_app.instance_path, 'school.db')
    
    if os.path.exists(src_path):
        try:
            shutil.copy2(src_path, db_path)
            flash(f"Database restored successfully from '{filename}'.", 'success')
        except Exception as e:
            flash(f"Error restoring backup: {e}", 'danger')
    else:
        flash('Backup file not found.', 'danger')
    return redirect(url_for('admin.backups'))

@admin.route('/backups/<path:filename>/delete', methods=['POST'])
@login_required
@admin_required
def delete_backup(filename):
    backup_dir = os.path.join(current_app.root_path, '..', 'backups')
    filepath = os.path.join(backup_dir, filename)
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
            flash(f"Backup file '{filename}' deleted.", 'success')
        except Exception as e:
            flash(f"Error deleting backup: {e}", 'danger')
    return redirect(url_for('admin.backups'))

# =========================================================================
# USER MANAGEMENT
# =========================================================================
ROLES = [
    SimpleNamespace(id=1, name='Super Administrator'),
    SimpleNamespace(id=2, name='Staff Member')
]

@admin.route('/users')
@login_required
@admin_required
def list_users():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin/users_list.html', users=users, roles=ROLES)

@admin.route('/users/create', methods=['POST'])
@login_required
@admin_required
def create_user():
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    role_id = request.form.get('role_id', type=int)

    if not username or not email or not password:
        flash('Please fill in all required fields.', 'danger')
        return redirect(url_for('admin.list_users'))

    existing = User.query.filter((User.username == username) | (User.email == email)).first()
    if existing:
        flash('A user with that username or email already exists.', 'danger')
        return redirect(url_for('admin.list_users'))

    is_admin = (role_id == 1)
    new_u = User(
        username=username,
        email=email,
        is_admin=is_admin,
        is_active=True
    )
    new_u.set_password(password)
    db.session.add(new_u)
    db.session.commit()
    flash(f"User '{username}' created successfully.", 'success')
    return redirect(url_for('admin.list_users'))

@admin.route('/users/<int:id>/update-role', methods=['POST'])
@login_required
@admin_required
def update_user_role(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('You cannot change your own role.', 'danger')
    else:
        role_id = request.form.get('role_id', type=int)
        user.is_admin = (role_id == 1)
        db.session.commit()
        flash(f"Role updated for '{user.username}'.", 'success')
    return redirect(url_for('admin.list_users'))

@admin.route('/users/<int:id>/toggle-status', methods=['POST'])
@login_required
@admin_required
def toggle_user_status(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('You cannot deactivate your own account.', 'danger')
    else:
        user.is_active = not user.is_active
        db.session.commit()
        status = 're-activated' if user.is_active else 'deactivated'
        flash(f"User '{user.username}' has been {status}.", 'success')
    return redirect(url_for('admin.list_users'))

@admin.route('/users/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
    else:
        db.session.delete(user)
        db.session.commit()
        flash(f"User '{user.username}' deleted successfully.", 'success')
    return redirect(url_for('admin.list_users'))


# =========================================================================
# SMS BROADCAST — PARENT CONTACTS & TWILIO INTEGRATION
# =========================================================================

def _clean_cell_value(val):
    """Safely extracts string content from an openpyxl cell value."""
    if val is None:
        return ''
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return f"{val:.0f}"
    elif isinstance(val, int):
        return str(val)
    return str(val).strip()


def _normalize_phone_number(raw_phone, default_country_code='+234'):
    """
    Normalizes raw phone numbers into standard E.164 format (+[country][number]).
    Handles Nigerian standard (080..., 070..., 090..., 081...), international (+...),
    and raw numeric strings from Excel.
    """
    if not raw_phone:
        return None

    cleaned = str(raw_phone).strip()
    if cleaned.lower() in ('none', 'null', 'n/a', '-'):
        return None

    # Handle float representation from Excel e.g. "2348012345678.0"
    if cleaned.endswith('.0'):
        cleaned = cleaned[:-2]

    # Remove all non-digit and non-plus characters
    has_plus = cleaned.startswith('+')
    digits_only = ''.join(c for c in cleaned if c.isdigit())

    if not digits_only:
        return None

    # Clean default country prefix
    country_prefix = default_country_code.strip() if default_country_code else '+234'
    country_digits = ''.join(c for c in country_prefix if c.isdigit())
    if not country_prefix.startswith('+'):
        country_prefix = '+' + country_digits

    # Case 1: Already has leading +
    if has_plus:
        return '+' + digits_only

    # Case 2: Starts with local zero (e.g. 08031234567 in Nigeria -> +2348031234567)
    if digits_only.startswith('0'):
        return f"{country_prefix}{digits_only[1:]}"

    # Case 3: Starts with country code digits directly without + (e.g. 2348031234567)
    if country_digits and digits_only.startswith(country_digits):
        return '+' + digits_only

    # Case 4: 10 digits without leading 0 (e.g. 8031234567) -> prepend default country code
    if len(digits_only) == 10:
        return f"{country_prefix}{digits_only}"

    # Default fallback: prepend default country prefix
    return f"{country_prefix}{digits_only}"


def _get_sms_setting(key, fallback_env=None):
    """Fetches setting from DB, falling back to environment variables / config."""
    setting = Setting.query.filter_by(key=key).first()
    if setting and setting.value and setting.value.strip():
        return setting.value.strip()
    if fallback_env:
        val = os.environ.get(fallback_env) or current_app.config.get(fallback_env)
        if val:
            return str(val).strip()
    return ''


def _get_twilio_client():
    """Initializes Twilio client with validated credentials."""
    account_sid = _get_sms_setting('twilio_account_sid', 'TWILIO_ACCOUNT_SID')
    auth_token = _get_sms_setting('twilio_auth_token', 'TWILIO_AUTH_TOKEN')
    from_number = _get_sms_setting('twilio_from_number', 'TWILIO_FROM_NUMBER') or _get_sms_setting('twilio_from_number', 'TWILIO_PHONE_NUMBER')
    default_country_code = _get_sms_setting('twilio_default_country_code') or '+234'

    if not account_sid or not auth_token or not from_number:
        missing = []
        if not account_sid:
            missing.append('Account SID')
        if not auth_token:
            missing.append('Auth Token')
        if not from_number:
            missing.append('From Number')
        return None, None, default_country_code, f"Missing Twilio configuration: {', '.join(missing)}."

    try:
        from twilio.rest import Client
    except ImportError:
        return None, None, default_country_code, "Twilio library is not installed on the server (pip install twilio)."

    try:
        client = Client(account_sid, auth_token)
        return client, from_number, default_country_code, None
    except Exception as exc:
        return None, None, default_country_code, f"Failed to initialize Twilio client: {exc}"


@admin.route('/sms')
@login_required
@admin_required
def sms_broadcast():
    from app.models.cms import ParentContact, SMSBroadcast
    contacts = ParentContact.query.order_by(ParentContact.name).all()
    broadcasts = SMSBroadcast.query.order_by(SMSBroadcast.created_at.desc()).limit(25).all()
    sid = _get_sms_setting('twilio_account_sid', 'TWILIO_ACCOUNT_SID')
    token = _get_sms_setting('twilio_auth_token', 'TWILIO_AUTH_TOKEN')
    from_num = _get_sms_setting('twilio_from_number', 'TWILIO_FROM_NUMBER') or _get_sms_setting('twilio_from_number', 'TWILIO_PHONE_NUMBER')
    country_code = _get_sms_setting('twilio_default_country_code') or '+234'
    twilio_configured = bool(sid and token and from_num)

    return render_template(
        'admin/sms_broadcast.html',
        contacts=contacts,
        broadcasts=broadcasts,
        twilio_configured=twilio_configured,
        twilio_account_sid=sid,
        twilio_auth_token=token,
        twilio_from_number=from_num,
        twilio_default_country_code=country_code
    )


@admin.route('/sms/save-config', methods=['POST'])
@login_required
@admin_required
def sms_save_config():
    for key in ['twilio_account_sid', 'twilio_auth_token', 'twilio_from_number', 'twilio_default_country_code']:
        value = request.form.get(key, '').strip()
        setting = Setting.query.filter_by(key=key).first()
        if setting:
            setting.value = value
        else:
            db.session.add(Setting(key=key, value=value))
    db.session.commit()
    flash('SMS configuration saved successfully.', 'success')
    return redirect(url_for('admin.sms_broadcast'))


@admin.route('/sms/test', methods=['POST'])
@login_required
@admin_required
def sms_send_test():
    test_phone = request.form.get('test_phone', '').strip()
    test_message = request.form.get('test_message', '').strip() or 'This is a test SMS from your school management portal.'

    if not test_phone:
        flash('Please enter a destination phone number to test.', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    client, from_number, default_country, err = _get_twilio_client()
    if err:
        flash(f'SMS Test Error: {err}', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    normalized_to = _normalize_phone_number(test_phone, default_country)
    if not normalized_to:
        flash(f'Invalid test phone number: {test_phone}', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    try:
        msg = client.messages.create(
            body=test_message,
            from_=from_number,
            to=normalized_to
        )
        flash(f'✅ Test SMS sent successfully to {normalized_to}! Message SID: {msg.sid}', 'success')
    except Exception as exc:
        current_app.logger.error(f"Twilio test SMS failed to {normalized_to}: {exc}")
        flash(f'❌ Twilio Test Failed: {exc}', 'danger')

    return redirect(url_for('admin.sms_broadcast'))


@admin.route('/sms/upload-contacts', methods=['POST'])
@login_required
@admin_required
def sms_upload_contacts():
    from app.models.cms import ParentContact
    try:
        import openpyxl
    except ImportError:
        flash('The openpyxl library is not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    file = request.files.get('excel_file')
    if not file or file.filename == '':
        flash('Please select an Excel file to upload.', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    ext = os.path.splitext(secure_filename(file.filename))[1].lower()
    if ext not in ('.xlsx', '.xls'):
        flash('Invalid file format. Please upload a .xlsx or .xls Excel spreadsheet.', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    default_country = _get_sms_setting('twilio_default_country_code') or '+234'

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # Auto-detect column positions from row 1 headers
        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            flash('The uploaded Excel file appears to be empty.', 'danger')
            return redirect(url_for('admin.sms_broadcast'))

        raw_headers = [_clean_cell_value(cell.value).lower() for cell in header_row]

        phone_col = name_col = grade_col = class_col = None
        for i, h in enumerate(raw_headers):
            if phone_col is None and any(k in h for k in ('phone', 'mobile', 'contact', 'tel', 'number', 'gsm')):
                phone_col = i
            elif name_col is None and any(k in h for k in ('name', 'parent', 'guardian', 'father', 'mother', 'fullname')):
                name_col = i
            elif grade_col is None and any(k in h for k in ('grade', 'level', 'stage', 'form')):
                grade_col = i
            elif class_col is None and any(k in h for k in ('class', 'section', 'arm', 'room')):
                class_col = i

        if phone_col is None:
            flash(
                'Could not locate a phone number column in the header row. '
                'Please ensure your Excel sheet has a header row with a column named '
                '"Phone", "Mobile", "Contact", "GSM", or "Number".',
                'danger'
            )
            return redirect(url_for('admin.sms_broadcast'))

        # Replace or append mode
        if request.form.get('upload_mode') == 'replace':
            ParentContact.query.delete()
            db.session.flush()

        added = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            def _cell(col):
                if col is not None and col < len(row):
                    return _clean_cell_value(row[col])
                return ''

            phone_raw = _cell(phone_col)
            normalized_phone = _normalize_phone_number(phone_raw, default_country)
            if not normalized_phone:
                skipped += 1
                continue

            name_val = _cell(name_col) or 'Parent'
            grade_val = _cell(grade_col)
            class_val = _cell(class_col)

            db.session.add(ParentContact(
                name=name_val,
                phone=normalized_phone,
                grade=grade_val,
                class_name=class_val,
            ))
            added += 1

        db.session.commit()
        flash(f'✅ Successfully imported {added} contact(s). {skipped} invalid/blank row(s) skipped.', 'success')

    except Exception as exc:
        db.session.rollback()
        current_app.logger.error(f"Error reading Excel contact file: {exc}")
        flash(f'Error reading Excel file: {exc}', 'danger')

    return redirect(url_for('admin.sms_broadcast'))


@admin.route('/sms/send', methods=['POST'])
@login_required
@admin_required
def sms_send_broadcast():
    from app.models.cms import ParentContact, SMSBroadcast

    message_text = request.form.get('message', '').strip()
    if not message_text:
        flash('SMS message body cannot be empty.', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    contacts = ParentContact.query.all()
    if not contacts:
        flash('No parent contacts found in database. Please upload an Excel contact file first.', 'warning')
        return redirect(url_for('admin.sms_broadcast'))

    client, from_number, default_country, err = _get_twilio_client()
    if err:
        flash(f'Twilio Error: {err}', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    sent = 0
    failed = 0
    error_samples = []

    for contact in contacts:
        target_phone = _normalize_phone_number(contact.phone, default_country) or contact.phone
        try:
            client.messages.create(
                body=message_text,
                from_=from_number,
                to=target_phone
            )
            sent += 1
        except Exception as exc:
            failed += 1
            err_msg = str(exc)
            current_app.logger.error(f"Twilio SMS broadcast error for {contact.name} ({target_phone}): {err_msg}")
            if len(error_samples) < 3:
                error_samples.append(f"{target_phone}: {err_msg}")

            # If it's a global credentials/auth error (e.g. HTTP 401), stop immediately to avoid hammering Twilio
            if 'Authenticate' in err_msg or '20003' in err_msg or '21212' in err_msg:
                break

    error_summary = " | ".join(error_samples) if error_samples else None

    broadcast = SMSBroadcast(
        message=message_text,
        total_recipients=len(contacts),
        sent_count=sent,
        failed_count=failed,
        sent_by=current_user.username if current_user and current_user.is_authenticated else 'Admin',
        error_details=error_summary
    )
    db.session.add(broadcast)
    db.session.commit()

    if failed == 0:
        flash(f'✅ SMS Broadcast complete! Successfully delivered to all {sent} contact(s).', 'success')
    elif sent > 0:
        flash(f'⚠️ SMS Broadcast finished with partial failures — Sent: {sent} | Failed: {failed}. Details: {error_summary or "Check logs."}', 'warning')
    else:
        flash(f'❌ SMS Broadcast failed for all recipients. Reason: {error_summary or "Check Twilio credentials and recipient formats."}', 'danger')

    return redirect(url_for('admin.sms_broadcast'))


@admin.route('/sms/send-single', methods=['POST'])
@login_required
@admin_required
def sms_send_single():
    from app.models.cms import SMSBroadcast

    phone = request.form.get('phone', '').strip()
    recipient_name = request.form.get('recipient_name', '').strip()
    message_text = request.form.get('message', '').strip()

    if not phone:
        flash('Please enter a recipient phone number.', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    if not message_text:
        flash('SMS message body cannot be empty.', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    client, from_number, default_country, err = _get_twilio_client()
    if err:
        flash(f'Twilio Configuration Error: {err}', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    normalized_to = _normalize_phone_number(phone, default_country)
    if not normalized_to:
        flash(f'Invalid phone number format: {phone}', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    target_label = f"{recipient_name} ({normalized_to})" if recipient_name else normalized_to

    try:
        msg = client.messages.create(
            body=message_text,
            from_=from_number,
            to=normalized_to
        )

        broadcast = SMSBroadcast(
            message=f"[Single to {target_label}]: {message_text}",
            total_recipients=1,
            sent_count=1,
            failed_count=0,
            sent_by=current_user.username if current_user and current_user.is_authenticated else 'Admin',
            error_details=None
        )
        db.session.add(broadcast)
        db.session.commit()

        flash(f'✅ SMS sent successfully to {target_label}! (Message SID: {msg.sid})', 'success')
    except Exception as exc:
        err_msg = str(exc)
        current_app.logger.error(f"Twilio single SMS error to {target_label}: {err_msg}")

        broadcast = SMSBroadcast(
            message=f"[Single to {target_label}]: {message_text}",
            total_recipients=1,
            sent_count=0,
            failed_count=1,
            sent_by=current_user.username if current_user and current_user.is_authenticated else 'Admin',
            error_details=err_msg
        )
        db.session.add(broadcast)
        db.session.commit()

        flash(f'❌ Failed to send SMS to {target_label}: {err_msg}', 'danger')

    return redirect(url_for('admin.sms_broadcast'))


@admin.route('/sms/add-contact', methods=['POST'])
@login_required
@admin_required
def sms_add_contact():
    from app.models.cms import ParentContact

    name = request.form.get('name', '').strip() or 'Parent'
    phone = request.form.get('phone', '').strip()
    grade = request.form.get('grade', '').strip()
    class_name = request.form.get('class_name', '').strip()

    if not phone:
        flash('Please enter a phone number.', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    default_country = _get_sms_setting('twilio_default_country_code') or '+234'
    normalized_phone = _normalize_phone_number(phone, default_country)
    if not normalized_phone:
        flash(f'Invalid phone number format: {phone}', 'danger')
        return redirect(url_for('admin.sms_broadcast'))

    contact = ParentContact(
        name=name,
        phone=normalized_phone,
        grade=grade,
        class_name=class_name
    )
    db.session.add(contact)
    db.session.commit()
    flash(f'✅ Contact "{name}" ({normalized_phone}) added successfully.', 'success')
    return redirect(url_for('admin.sms_broadcast'))


@admin.route('/sms/delete-contact/<int:id>', methods=['POST'])
@login_required
@admin_required
def sms_delete_contact(id):
    from app.models.cms import ParentContact
    contact = ParentContact.query.get_or_404(id)
    name = contact.name
    db.session.delete(contact)
    db.session.commit()
    flash(f'Contact "{name}" removed.', 'info')
    return redirect(url_for('admin.sms_broadcast'))


@admin.route('/sms/clear-contacts', methods=['POST'])
@login_required
@admin_required
def sms_clear_contacts():
    from app.models.cms import ParentContact
    count = ParentContact.query.count()
    ParentContact.query.delete()
    db.session.commit()
    flash(f'Cleared all {count} parent contact(s).', 'info')
    return redirect(url_for('admin.sms_broadcast'))


@admin.route('/sms/delete-broadcast/<int:id>', methods=['POST'])
@login_required
@admin_required
def sms_delete_broadcast(id):
    from app.models.cms import SMSBroadcast
    broadcast = SMSBroadcast.query.get_or_404(id)
    db.session.delete(broadcast)
    db.session.commit()
    flash('Broadcast log record deleted.', 'info')
    return redirect(url_for('admin.sms_broadcast'))


@admin.route('/sms/clear-broadcasts', methods=['POST'])
@login_required
@admin_required
def sms_clear_broadcasts():
    from app.models.cms import SMSBroadcast
    count = SMSBroadcast.query.count()
    SMSBroadcast.query.delete()
    db.session.commit()
    flash(f'Cleared all {count} broadcast log record(s).', 'info')
    return redirect(url_for('admin.sms_broadcast'))




